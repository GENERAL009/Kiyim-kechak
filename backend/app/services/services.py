from sqlalchemy.orm import Session
from sqlalchemy import func
from fastapi import HTTPException, status
from typing import List, Optional, Tuple
import random
import io
from datetime import datetime

# ReportLab & OpenPyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from app.models.models import (
    User, Role, Category, Brand, Product, ProductVariant,
    Warehouse, Inventory, InventoryTransaction, Customer, Order, OrderItem, AuditLog
)
from app.schemas.schemas import (
    UserCreate, UserUpdate, ProductCreate, ProductVariantCreate,
    InventoryTransactionCreate, OrderCreate, OrderItemCreate
)
from app.repositories import (
    user_repo, role_repo, product_repo, product_variant_repo,
    warehouse_repo, inventory_repo, transaction_repo, customer_repo, order_repo, audit_log_repo
)
from app.core.security import hash_password, verify_password

# ==========================================
# AUDIT LOG SERVICE
# ==========================================
class AuditLogService:
    @staticmethod
    def log(db: Session, user_id: int, action: str, details: Optional[str] = None) -> AuditLog:
        db_log = AuditLog(
            user_id=user_id,
            action=action,
            details=details
        )
        db.add(db_log)
        db.commit()
        db.refresh(db_log)
        return db_log

# ==========================================
# AUTH SERVICE
# ==========================================
class AuthService:
    @staticmethod
    def authenticate_user(db: Session, email: str, password: str) -> Optional[User]:
        user = user_repo.get_by_email(db, email)
        if not user:
            return None
        if not verify_password(password, user.hashed_password):
            return None
        if not user.is_active:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="User account is inactive"
            )
        return user

    @staticmethod
    def create_user(db: Session, user_in: UserCreate, creator_id: Optional[int] = None) -> User:
        existing_user = user_repo.get_by_email(db, user_in.email)
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Email already registered"
            )
        
        role = role_repo.get(db, user_in.role_id)
        if not role:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Role does not exist"
            )

        hashed = hash_password(user_in.password)
        db_user = User(
            email=user_in.email,
            hashed_password=hashed,
            full_name=user_in.full_name,
            role_id=user_in.role_id,
            is_active=user_in.is_active
        )
        db.add(db_user)
        db.commit()
        db.refresh(db_user)

        if creator_id:
            AuditLogService.log(
                db, 
                user_id=creator_id, 
                action="CREATE_USER", 
                details=f"Created user {db_user.email} with role {role.name}"
            )

        return db_user

# ==========================================
# PRODUCT SERVICE
# ==========================================
class ProductService:
    @staticmethod
    def generate_sku(db: Session, category_id: int, brand_id: int, size: str, color: str) -> str:
        # PRD-CAT-BRD-SIZE-COLOR-RAND
        cat = db.query(Category).get(category_id)
        brd = db.query(Brand).get(brand_id)
        cat_code = cat.name[:3].upper() if cat else "GEN"
        brd_code = brd.name[:3].upper() if brd else "GEN"
        size_code = size.upper()
        color_code = color[:3].upper()
        
        while True:
            rand_num = random.randint(1000, 9999)
            sku = f"WMS-{cat_code}-{brd_code}-{size_code}-{color_code}-{rand_num}"
            if not product_variant_repo.get_by_sku(db, sku):
                return sku

    @staticmethod
    def generate_barcode(db: Session) -> str:
        while True:
            # 12-digit barcode starting with 99
            barcode = "99" + "".join([str(random.randint(0, 9)) for _ in range(10)])
            if not product_variant_repo.get_by_barcode(db, barcode):
                return barcode

    @staticmethod
    def create_product(db: Session, product_in: ProductCreate, user_id: int) -> Product:
        # Check category and brand
        category = db.query(Category).get(product_in.category_id)
        brand = db.query(Brand).get(product_in.brand_id)
        if not category or not brand:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Category or Brand not found"
            )

        # Create main product
        db_product = Product(
            name=product_in.name,
            description=product_in.description,
            category_id=product_in.category_id,
            brand_id=product_in.brand_id
        )
        db.add(db_product)
        db.commit()
        db.refresh(db_product)

        # Create variants
        for var in product_in.variants:
            sku = var.sku if var.sku else ProductService.generate_sku(db, product_in.category_id, product_in.brand_id, var.size, var.color)
            barcode = var.barcode if var.barcode else ProductService.generate_barcode(db)

            # Check duplication
            if product_variant_repo.get_by_sku(db, sku):
                raise HTTPException(status_code=400, detail=f"SKU {sku} already exists")
            if product_variant_repo.get_by_barcode(db, barcode):
                raise HTTPException(status_code=400, detail=f"Barcode {barcode} already exists")

            db_var = ProductVariant(
                product_id=db_product.id,
                sku=sku,
                barcode=barcode,
                size=var.size,
                color=var.color,
                price=var.price,
                image_url=var.image_url
            )
            db.add(db_var)
        
        db.commit()
        db.refresh(db_product)

        AuditLogService.log(
            db,
            user_id=user_id,
            action="CREATE_PRODUCT",
            details=f"Created product {db_product.name} with {len(db_product.variants)} variants"
        )
        return db_product

# ==========================================
# INVENTORY SERVICE
# ==========================================
class InventoryService:
    @staticmethod
    def add_stock(db: Session, tx_in: InventoryTransactionCreate, user_id: int) -> InventoryTransaction:
        if not tx_in.to_warehouse_id:
            raise HTTPException(status_code=400, detail="Destination warehouse is required for adding stock")
        
        # Verify variant and warehouse
        variant = product_variant_repo.get(db, tx_in.variant_id)
        warehouse = warehouse_repo.get(db, tx_in.to_warehouse_id)
        if not variant or not warehouse:
            raise HTTPException(status_code=404, detail="Variant or Warehouse not found")

        # Get or create inventory item
        inv = inventory_repo.get_by_warehouse_and_variant(db, tx_in.to_warehouse_id, tx_in.variant_id)
        if not inv:
            inv = Inventory(
                warehouse_id=tx_in.to_warehouse_id,
                variant_id=tx_in.variant_id,
                quantity=0
            )
            db.add(inv)
        
        inv.quantity += tx_in.quantity
        db.commit()

        # Log transaction
        tx = InventoryTransaction(
            transaction_type="INCOMING",
            variant_id=tx_in.variant_id,
            quantity=tx_in.quantity,
            to_warehouse_id=tx_in.to_warehouse_id,
            user_id=user_id,
            remarks=tx_in.remarks
        )
        db.add(tx)
        db.commit()
        db.refresh(tx)

        AuditLogService.log(
            db, user_id=user_id, action="STOCK_IN",
            details=f"Added {tx_in.quantity} items of SKU {variant.sku} to Warehouse {warehouse.name}"
        )
        return tx

    @staticmethod
    def remove_stock(db: Session, tx_in: InventoryTransactionCreate, user_id: int) -> InventoryTransaction:
        if not tx_in.from_warehouse_id:
            raise HTTPException(status_code=400, detail="Source warehouse is required for removing stock")

        variant = product_variant_repo.get(db, tx_in.variant_id)
        warehouse = warehouse_repo.get(db, tx_in.from_warehouse_id)
        if not variant or not warehouse:
            raise HTTPException(status_code=404, detail="Variant or Warehouse not found")

        inv = inventory_repo.get_by_warehouse_and_variant(db, tx_in.from_warehouse_id, tx_in.variant_id)
        if not inv or inv.quantity < tx_in.quantity:
            raise HTTPException(status_code=400, detail="Insufficient stock in the warehouse")

        inv.quantity -= tx_in.quantity
        db.commit()

        # Log transaction
        tx = InventoryTransaction(
            transaction_type="OUTGOING",
            variant_id=tx_in.variant_id,
            quantity=tx_in.quantity,
            from_warehouse_id=tx_in.from_warehouse_id,
            user_id=user_id,
            remarks=tx_in.remarks
        )
        db.add(tx)
        db.commit()
        db.refresh(tx)

        AuditLogService.log(
            db, user_id=user_id, action="STOCK_OUT",
            details=f"Removed {tx_in.quantity} items of SKU {variant.sku} from Warehouse {warehouse.name}"
        )
        return tx

    @staticmethod
    def transfer_stock(db: Session, tx_in: InventoryTransactionCreate, user_id: int) -> InventoryTransaction:
        if not tx_in.from_warehouse_id or not tx_in.to_warehouse_id:
            raise HTTPException(status_code=400, detail="Both source and destination warehouses are required for transfer")

        if tx_in.from_warehouse_id == tx_in.to_warehouse_id:
            raise HTTPException(status_code=400, detail="Source and destination warehouses cannot be the same")

        variant = product_variant_repo.get(db, tx_in.variant_id)
        from_w = warehouse_repo.get(db, tx_in.from_warehouse_id)
        to_w = warehouse_repo.get(db, tx_in.to_warehouse_id)
        if not variant or not from_w or not to_w:
            raise HTTPException(status_code=404, detail="Variant or Warehouses not found")

        # Deduct from source
        inv_from = inventory_repo.get_by_warehouse_and_variant(db, tx_in.from_warehouse_id, tx_in.variant_id)
        if not inv_from or inv_from.quantity < tx_in.quantity:
            raise HTTPException(status_code=400, detail="Insufficient stock in source warehouse")

        # Add to destination
        inv_to = inventory_repo.get_by_warehouse_and_variant(db, tx_in.to_warehouse_id, tx_in.variant_id)
        if not inv_to:
            inv_to = Inventory(
                warehouse_id=tx_in.to_warehouse_id,
                variant_id=tx_in.variant_id,
                quantity=0
            )
            db.add(inv_to)
        
        inv_from.quantity -= tx_in.quantity
        inv_to.quantity += tx_in.quantity
        db.commit()

        # Log transaction
        tx = InventoryTransaction(
            transaction_type="TRANSFER",
            variant_id=tx_in.variant_id,
            quantity=tx_in.quantity,
            from_warehouse_id=tx_in.from_warehouse_id,
            to_warehouse_id=tx_in.to_warehouse_id,
            user_id=user_id,
            remarks=tx_in.remarks
        )
        db.add(tx)
        db.commit()
        db.refresh(tx)

        AuditLogService.log(
            db, user_id=user_id, action="STOCK_TRANSFER",
            details=f"Transferred {tx_in.quantity} items of SKU {variant.sku} from {from_w.name} to {to_w.name}"
        )
        return tx

    @staticmethod
    def adjust_stock(db: Session, tx_in: InventoryTransactionCreate, user_id: int) -> InventoryTransaction:
        # Stock adjustment sets or modifies inventory levels directly
        if not tx_in.to_warehouse_id:
            raise HTTPException(status_code=400, detail="Warehouse is required for adjustment")

        variant = product_variant_repo.get(db, tx_in.variant_id)
        warehouse = warehouse_repo.get(db, tx_in.to_warehouse_id)
        if not variant or not warehouse:
            raise HTTPException(status_code=404, detail="Variant or Warehouse not found")

        inv = inventory_repo.get_by_warehouse_and_variant(db, tx_in.to_warehouse_id, tx_in.variant_id)
        if not inv:
            inv = Inventory(
                warehouse_id=tx_in.to_warehouse_id,
                variant_id=tx_in.variant_id,
                quantity=0
            )
            db.add(inv)

        old_qty = inv.quantity
        inv.quantity = tx_in.quantity  # In adjustment, quantity is the new target amount
        difference = inv.quantity - old_qty
        db.commit()

        tx = InventoryTransaction(
            transaction_type="ADJUSTMENT",
            variant_id=tx_in.variant_id,
            quantity=difference,  # Log the difference
            to_warehouse_id=tx_in.to_warehouse_id,
            user_id=user_id,
            remarks=f"Adjusted from {old_qty} to {tx_in.quantity}. " + (tx_in.remarks or "")
        )
        db.add(tx)
        db.commit()
        db.refresh(tx)

        AuditLogService.log(
            db, user_id=user_id, action="STOCK_ADJUST",
            details=f"Adjusted SKU {variant.sku} in {warehouse.name} from {old_qty} to {tx_in.quantity}"
        )
        return tx

# ==========================================
# SALES SERVICE
# ==========================================
class SalesService:
    @staticmethod
    def create_order(db: Session, order_in: OrderCreate, seller_id: int) -> Order:
        if not order_in.items:
            raise HTTPException(status_code=400, detail="Order items cannot be empty")

        # Get or create customer
        customer = None
        if order_in.customer_id:
            customer = customer_repo.get(db, order_in.customer_id)
        elif order_in.customer_phone:
            customer = customer_repo.get_by_phone(db, order_in.customer_phone)
            if not customer:
                customer = Customer(
                    full_name=order_in.customer_name or "Walk-in Customer",
                    phone=order_in.customer_phone,
                    email=None
                )
                db.add(customer)
                db.commit()
                db.refresh(customer)
        
        if not customer:
            # Default Walk-in Customer
            customer = customer_repo.get_by_phone(db, "999999999")
            if not customer:
                customer = Customer(
                    full_name="Walk-in Customer",
                    phone="999999999",
                    email=None
                )
                db.add(customer)
                db.commit()
                db.refresh(customer)

        total_amount = 0.0
        order_items_to_create = []

        # Process each item and deduct stock
        for item in order_in.items:
            variant = product_variant_repo.get(db, item.variant_id)
            if not variant:
                raise HTTPException(status_code=404, detail=f"Variant with ID {item.variant_id} not found")

            # Check stock across all warehouses, deduct from first available
            qty_needed = item.quantity
            inventory_records = db.query(Inventory).filter(
                Inventory.variant_id == item.variant_id,
                Inventory.quantity > 0
            ).all()

            total_available = sum([r.quantity for r in inventory_records])
            if total_available < qty_needed:
                raise HTTPException(
                    status_code=400,
                    detail=f"Insufficient total stock for {variant.sku}. Available: {total_available}, Requested: {qty_needed}"
                )

            # Deduct stock
            for inv in inventory_records:
                if qty_needed <= 0:
                    break
                deduct_qty = min(inv.quantity, qty_needed)
                inv.quantity -= deduct_qty
                qty_needed -= deduct_qty

                # Log Transaction
                tx = InventoryTransaction(
                    transaction_type="OUTGOING",
                    variant_id=item.variant_id,
                    quantity=deduct_qty,
                    from_warehouse_id=inv.warehouse_id,
                    user_id=seller_id,
                    remarks=f"Sale. Order item processing."
                )
                db.add(tx)

            total_amount += variant.price * item.quantity
            order_item = OrderItem(
                variant_id=item.variant_id,
                quantity=item.quantity,
                unit_price=variant.price
            )
            order_items_to_create.append(order_item)

        # Create Order
        db_order = Order(
            customer_id=customer.id,
            seller_id=seller_id,
            total_amount=total_amount,
            status="COMPLETED"
        )
        db.add(db_order)
        db.commit()
        db.refresh(db_order)

        # Link order items
        for order_item in order_items_to_create:
            order_item.order_id = db_order.id
            db.add(order_item)
        
        db.commit()
        db.refresh(db_order)

        AuditLogService.log(
            db, user_id=seller_id, action="CREATE_ORDER",
            details=f"Created Order #{db_order.id} for Customer {customer.full_name}, Total: {total_amount}"
        )
        return db_order

# ==========================================
# REPORT SERVICE (Excel & PDF)
# ==========================================
class ReportService:
    @staticmethod
    def generate_sales_pdf(order: Order) -> bytes:
        # Create bytes buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'InvoiceTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1E293B'),
            spaceAfter=15
        )
        subtitle_style = ParagraphStyle(
            'InvoiceSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#64748B'),
            spaceAfter=5
        )
        bold_style = ParagraphStyle(
            'InvoiceBold',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#0F172A'),
            fontName='Helvetica-Bold'
        )

        elements = []

        # Title
        elements.append(Paragraph("CLOTHING STORE WMS", title_style))
        elements.append(Paragraph(f"INVOICE / SOTUV CHEKI #{order.id}", bold_style))
        elements.append(Paragraph(f"Sana: {order.created_at.strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
        elements.append(Paragraph(f"Sotuvchi: {order.seller.full_name}", subtitle_style))
        elements.append(Paragraph(f"Mijoz: {order.customer.full_name} ({order.customer.phone})", subtitle_style))
        elements.append(Spacer(1, 20))

        # Table header
        data = [["#", "Mahsulot SKU", "Rang/O'lcham", "Miqdor", "Dona narxi ($)", "Jami narx ($)"]]
        for idx, item in enumerate(order.items, start=1):
            var = item.variant
            prod_name = var.product.name if var.product else "Mahsulot"
            desc = f"{prod_name} ({var.size} / {var.color})"
            row = [
                str(idx),
                var.sku,
                desc,
                str(item.quantity),
                f"${item.unit_price:.2f}",
                f"${(item.quantity * item.unit_price):.2f}"
            ]
            data.append(row)

        # Totals row
        data.append(["", "", "", "", "Umumiy Summa:", f"${order.total_amount:.2f}"])

        # Table style
        t = Table(data, colWidths=[20, 140, 180, 50, 80, 80])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F1F5F9')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#0F172A')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('ALIGN', (3,1), (-1,-1), 'RIGHT'),
            ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#CBD5E1')),
            ('FONTNAME', (4,-1), (5,-1), 'Helvetica-Bold'),
            ('LINEABOVE', (4,-1), (5,-1), 1, colors.HexColor('#0F172A')),
            ('TOPPADDING', (0,-1), (-1,-1), 10),
        ]))
        
        elements.append(t)
        
        # Build document
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_inventory_excel(db: Session) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ombor Qoldiqlari"

        # Headers
        headers = ["Ombor Nomi", "Mahsulot Nomi", "SKU", "Shtrixkod", "O'lcham", "Rang", "Narxi ($)", "Miqdori", "Minimal Qoldiq"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data
        inventory_items = db.query(Inventory).all()
        for idx, item in enumerate(inventory_items, 2):
            var = item.variant
            prod = var.product
            ws.append([
                item.warehouse.name,
                prod.name,
                var.sku,
                var.barcode,
                var.size,
                var.color,
                var.price,
                item.quantity,
                item.min_stock_level
            ])

        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        # Save workbook to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
